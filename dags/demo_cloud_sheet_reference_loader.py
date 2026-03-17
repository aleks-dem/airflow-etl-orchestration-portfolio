from __future__ import annotations

import json
import os
import warnings
from datetime import datetime, timedelta
from typing import List, Tuple

import numpy as np
import pandas as pd
from airflow.decorators import dag, task
from airflow.models import Variable
from openpyxl import load_workbook
from zoneinfo import ZoneInfo

from utils import connectors as con
from utils import messages as msg

warnings.filterwarnings("ignore")


def get_params(key: str):
    params = {
        "schema_name": "analytics_demo",
        "main_table_name": "reference_snapshot",
        "insert_batch_size": 1_000_000,
        "google_folder_id_var": "demo_reference_google_folder_id",
        "sheet_name_var": "demo_reference_sheet_name",
        "directory": "dags",
    }
    return params[key]


def list_spreadsheets_in_folder(drive_service, folder_id: str):
    files = []
    page_token = None
    query = (
        f"'{folder_id}' in parents and "
        "mimeType='application/vnd.google-apps.spreadsheet' and trashed=false"
    )
    while True:
        resp = (
            drive_service.files()
            .list(
                q=query,
                fields="nextPageToken, files(id, name)",
                includeItemsFromAllDrives=True,
                supportsAllDrives=True,
                corpora="allDrives",
                pageToken=page_token,
            )
            .execute()
        )
        files.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return files


def read_spreadsheet(
    sheets_service,
    file_id: str,
    file_name: str,
    snapshot_dt: datetime,
    add_source_column: bool = True,
    source_column_name: str = "source_file",
    sheet_name: str | None = None,
    sheet_index: int | None = None,
) -> pd.DataFrame:
    meta = (
        sheets_service.spreadsheets()
        .get(spreadsheetId=file_id, fields="sheets(properties/title)")
        .execute()
    )
    sheets = meta.get("sheets", [])
    titles = [sheet["properties"]["title"] for sheet in sheets]

    if sheet_name is not None:
        if sheet_name not in titles:
            raise ValueError(f"Worksheet '{sheet_name}' not found in file '{file_name}'.")
        target_sheet = sheet_name
    else:
        idx = sheet_index if sheet_index is not None else 0
        if idx < 0 or idx >= len(titles):
            raise IndexError(f"Worksheet index {idx} is out of range for '{file_name}'.")
        target_sheet = titles[idx]

    result = (
        sheets_service.spreadsheets()
        .values()
        .get(spreadsheetId=file_id, range=target_sheet)
        .execute()
    )
    values = result.get("values", [])
    if not values:
        return pd.DataFrame()

    header = values[0]
    data_rows = values[1:]

    consecutive_blanks = 0
    cols_of_interest = []
    for idx, col_name in enumerate(header):
        if col_name is None or str(col_name).strip() == "":
            consecutive_blanks += 1
            if consecutive_blanks >= 3:
                break
        else:
            consecutive_blanks = 0
            cols_of_interest.append(idx)

    cols = [header[i] for i in cols_of_interest]
    filtered_data = []
    for row in data_rows:
        new_row = []
        for col_idx in cols_of_interest:
            if col_idx < len(row) and str(row[col_idx]).strip():
                new_row.append(row[col_idx])
            else:
                new_row.append(np.nan)
        filtered_data.append(new_row)

    df = pd.DataFrame(filtered_data, columns=cols)
    df.replace("", np.nan, inplace=True)
    df.replace(r"^\s*$", np.nan, regex=True, inplace=True)
    df = df.where(df.isna(), df.astype(str))

    if add_source_column:
        df[source_column_name] = file_name
    df["snapshot_dt"] = snapshot_dt
    return df


def standardize_df(df: pd.DataFrame):
    rename_map = {
        "ID": "record_id",
        "Name": "record_name",
        "Category": "category",
        "Type": "type",
        "Group": "group_name",
        "Segment": "segment",
        "Region": "region",
        "Channel": "channel",
        "Value": "metric_value",
        "Currency": "metric_currency",
        "Status": "status_tag",
        "Priority": "priority",
        "Owner": "owner_name",
        "Comments": "comments",
        "Date": "event_date",
        "Actuality": "actuality",
        "Tags": "tags",
    }
    df = df.rename(columns=lambda x: x.strip())
    return df.rename(columns=rename_map)


def load_expected_columns(table=None) -> List[str]:
    client = con.get_clickhouse_client()
    try:
        db = get_params("schema_name")
        table_name = table or get_params("main_table_name")
        query = f"""
            SELECT name
            FROM system.columns
            WHERE database = '{db}'
              AND table = '{table_name}'
            ORDER BY position
        """
        rows = client.query_df(query)
        return rows["name"].to_list()
    finally:
        client.close()


def validate_structure(
    df: pd.DataFrame,
    expected: List[str],
) -> Tuple[bool, List[str], List[str], List[str]]:
    dup_mask = df.columns.duplicated(keep=False)
    duplicates = list({col for col, dup in zip(df.columns, dup_mask) if dup})
    new_cols = list(set(df.columns) - set(expected))
    missing_cols = list(set(expected) - set(df.columns))
    is_valid = not (duplicates or new_cols)
    return is_valid, duplicates, new_cols, missing_cols


@dag(
    "demo_cloud_sheet_reference_loader",
    description="Cloud spreadsheet ingestion DAG with schema validation and quality reporting",
    schedule_interval="30 8 * * *",
    catchup=False,
    tags=["demo", "cloud-drive", "spreadsheets", "validation", "clickhouse"],
    default_args={
        "owner": "portfolio_demo",
        "start_date": datetime(2025, 1, 1),
        "retries": 1,
        "retry_delay": timedelta(minutes=10),
    },
    max_active_runs=1,
    params={
        "receiver_channel_id": "",
        "owner_channel_id": "",
        "report_warning_files": False,
        "report_bad_files": True,
        "report_duplicates": False,
    },
)
def demo_cloud_sheet_reference_loader():
    @task()
    def transfer_reference_data(**context):
        task_instance = context["task_instance"]
        folder_id = Variable.get(get_params("google_folder_id_var"))
        if not folder_id:
            raise ValueError("Airflow Variable demo_reference_google_folder_id is not set.")

        report_sheet_name = Variable.get(get_params("sheet_name_var"), default_var="Data")
        snapshot_dt = datetime.now(tz=ZoneInfo("UTC")).replace(microsecond=0)
        current_directory = Variable.get("etl_directory")
        dag_directory = f"{current_directory}/{get_params('directory')}"

        try:
            expected_columns = load_expected_columns()
            drive_service, sheets_service = con.get_google_services(dag_directory=dag_directory)
            files = list_spreadsheets_in_folder(drive_service, folder_id)

            good_dfs = {}
            bad_files = []
            warning_files = []

            for file_meta in files:
                print(f"Loading file: {file_meta['name']}")
                df = read_spreadsheet(
                    sheets_service,
                    file_meta["id"],
                    file_meta["name"],
                    snapshot_dt,
                    sheet_name=report_sheet_name,
                )
                if df.empty:
                    bad_files.append((file_meta["name"], "Empty sheet"))
                    continue

                df = standardize_df(df)
                if "record_name" not in df.columns:
                    bad_files.append((file_meta["name"], "Missing required column: record_name"))
                    continue
                df = df.loc[df["record_name"].notna(), :]

                ok, duplicates, unexpected, missing = validate_structure(df, expected_columns)
                if not ok:
                    bad_files.append(
                        (
                            file_meta["name"],
                            f"Unexpected structure: duplicates={duplicates}, unexpected={unexpected}",
                        )
                    )
                    continue

                if missing:
                    warning_files.append(
                        (
                            file_meta["name"],
                            f"Missing optional columns, will be filled with NaN: {missing}",
                        )
                    )

                for col in expected_columns:
                    if col not in df.columns:
                        df[col] = np.nan
                df = df[expected_columns]
                good_dfs[file_meta["name"]] = df

            if not good_dfs:
                raise ValueError(
                    f"No valid spreadsheets were found in folder '{folder_id}' for loading into ClickHouse."
                )

            consolidated = pd.concat(good_dfs.values(), ignore_index=True, sort=False)
            key_column = "record_id"

            duplicate_mask = consolidated[key_column].duplicated(keep=False)
            duplicates_df = consolidated[duplicate_mask].copy()
            deduped_df = consolidated.drop_duplicates(subset=[key_column], keep="last")

            raw_record_id = deduped_df[key_column].astype("string")
            clean_record_id = raw_record_id.str.strip()
            broken_mask = clean_record_id.isna() | (clean_record_id == "") | (~clean_record_id.str.contains("_", na=False))

            record_name = deduped_df["record_name"].astype("string").fillna("").str.strip()
            category = deduped_df.get("category", pd.Series(dtype="string")).astype("string").fillna("").str.strip()
            actuality = deduped_df.get("actuality", pd.Series(dtype="string")).astype("string").fillna("").str.strip()

            generated_record_id = (
                (record_name + "_" + category + "_" + actuality)
                .str.replace(r"__+", "_", regex=True)
                .str.strip("_")
            )
            deduped_df.loc[broken_mask, key_column] = generated_record_id[broken_mask]
            deduped_df[key_column] = deduped_df[key_column].fillna("").astype("string")

            timezone_columns = duplicates_df.select_dtypes(include=["datetimetz"]).columns
            for col in timezone_columns:
                duplicates_df[col] = duplicates_df[col].dt.tz_localize(None)
            duplicate_records = json.loads(duplicates_df.to_json(orient="records", date_format="iso"))

            task_instance.xcom_push(key="bad_files", value=bad_files)
            task_instance.xcom_push(key="warn_files", value=warning_files)
            task_instance.xcom_push(key="duplicates", value=duplicate_records)

            con.insert_df_to_dwh(
                deduped_df,
                get_params("schema_name"),
                get_params("main_table_name"),
                get_params("insert_batch_size"),
                trunc=False,
            )
        except Exception as exc:
            msg.notify_on_failure(context, exc)
            raise

    @task(trigger_rule="all_done")
    def notification(**context):
        try:
            task_instance = context["task_instance"]
            bad_files = task_instance.xcom_pull(
                task_ids="transfer_reference_data",
                key="bad_files",
            ) or []
            warning_files = task_instance.xcom_pull(
                task_ids="transfer_reference_data",
                key="warn_files",
            ) or []
            duplicate_records = task_instance.xcom_pull(
                task_ids="transfer_reference_data",
                key="duplicates",
            ) or []

            report_bad_files = context["params"]["report_bad_files"]
            report_warning_files = context["params"]["report_warning_files"]
            report_duplicates = context["params"]["report_duplicates"]

            if not (
                (report_bad_files and bad_files)
                or (report_warning_files and warning_files)
                or (report_duplicates and duplicate_records)
            ):
                return

            current_directory = Variable.get("etl_directory")
            report_dir = f"{current_directory}/{get_params('directory')}"
            report_path = os.path.join(
                report_dir,
                f"reference_loading_issues_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            )

            with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
                pd.DataFrame(bad_files, columns=["file", "error"]).to_excel(
                    writer, sheet_name="bad_files", index=False
                )
                pd.DataFrame(warning_files, columns=["file", "warning"]).to_excel(
                    writer, sheet_name="warning_files", index=False
                )
                if duplicate_records:
                    pd.DataFrame(duplicate_records).to_excel(
                        writer, sheet_name="duplicates", index=False
                    )

            workbook = load_workbook(report_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for col in sheet.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        value = str(cell.value or "")
                        if len(value) > max_length:
                            max_length = len(value)
                    sheet.column_dimensions[col_letter].width = min(max_length + 2, 100)
            workbook.save(report_path)

            channel_id = context["params"].get("receiver_channel_id")
            if channel_id:
                msg.send_team_chat_message(
                    message=(
                        "#### Data quality issues were found during spreadsheet ingestion. "
                        "Please review the attached report."
                    ),
                    channel_id=channel_id,
                    file_path=report_path,
                )
            os.remove(report_path)
        except Exception as exc:
            msg.notify_on_failure(context, exc)
            raise

    main_task = transfer_reference_data()
    notify_task = notification()
    main_task >> notify_task


demo_cloud_sheet_reference_loader()
